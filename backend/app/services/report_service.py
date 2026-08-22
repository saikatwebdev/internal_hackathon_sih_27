import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session
from app.database.models import ClassSession, Student, Enrollment, Attendance, Subject, Branch, Faculty


class ReportService:
    @staticmethod
    def get_session_attendance_matrix(db: Session, class_session_id: str):
        """
        Retrieves the complete attendance matrix for a class session.
        Calculates population as: All enrolled students LEFT JOIN attendance.
        """
        session = db.query(ClassSession).filter(ClassSession.id == class_session_id).first()
        if not session:
            return None, []

        # Fetch enrolled students in session's branch, year, semester, section
        students = (
            db.query(Student)
            .join(Enrollment, Student.id == Enrollment.student_id)
            .filter(
                Enrollment.branch_id == session.branch_id,
                Enrollment.year == session.year,
                Enrollment.semester == session.semester,
                Enrollment.section == session.section,
                Enrollment.status == "active",
            )
            .order_by(Student.roll_no.asc())
            .all()
        )

        # Fetch attendance records for this session
        attendances = (
            db.query(Attendance)
            .filter(Attendance.class_session_id == class_session_id)
            .all()
        )
        att_map = {att.student_id: att for att in attendances}

        records = []
        for stud in students:
            att = att_map.get(stud.id)
            if att:
                entry_str = att.entry_time.strftime("%H:%M:%S") if att.entry_time else "--"
                exit_str = att.exit_time.strftime("%H:%M:%S") if att.exit_time else "--"
                duration = att.duration_minutes if att.duration_minutes is not None else "--"
                qr_v = "YES" if att.qr_verified else "NO"
                face_v = "YES" if att.face_verified else "NO"
                status = att.status.upper() if att.status else "ABSENT"
            else:
                entry_str = "--"
                exit_str = "--"
                duration = "--"
                qr_v = "NO"
                face_v = "NO"
                status = "ABSENT"

            records.append({
                "student_id": stud.id,
                "roll_no": stud.roll_no,
                "student_name": stud.name,
                "phone": stud.phone,
                "email": stud.email or "",
                "entry_time": entry_str,
                "exit_time": exit_str,
                "duration_minutes": duration,
                "qr_verified": qr_v,
                "face_verified": face_v,
                "status": status,
            })

        return session, records

    @staticmethod
    def export_session_csv(db: Session, class_session_id: str) -> str:
        session, records = ReportService.get_session_attendance_matrix(db, class_session_id)
        if not records:
            return ""

        df = pd.DataFrame(records)
        df = df.drop(columns=["student_id"])
        df.columns = [
            "Roll No", "Student Name", "Phone", "Email",
            "Entry Time", "Exit Time", "Duration (min)",
            "QR Verified", "Face Verified", "Status"
        ]
        return df.to_csv(index=False)

    @staticmethod
    def export_session_excel(db: Session, class_session_id: str) -> bytes:
        session, records = ReportService.get_session_attendance_matrix(db, class_session_id)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Class Attendance"

        # Styles
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        title_font = Font(name="Calibri", size=14, bold=True, color="1E3A8A")
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        # Title Block
        if session:
            subj_name = session.subject.subject_name if session.subject else "N/A"
            subj_code = session.subject.subject_code if session.subject else "N/A"
            ws.cell(row=1, column=1, value=f"Attendance Report: {subj_code} - {subj_name}").font = title_font
            ws.cell(row=2, column=1, value=f"Date: {session.class_date} | Time: {session.start_time} - {session.end_time} | Section: {session.section}")

        headers = [
            "Roll No", "Student Name", "Phone", "Email",
            "Entry Time", "Exit Time", "Duration (min)",
            "QR Verified", "Face Verified", "Status"
        ]

        start_row = 4
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num, value=header_title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        present_count = 0
        absent_count = 0

        for r_idx, rec in enumerate(records, start_row + 1):
            status = rec["status"]
            if "PRESENT" in status:
                present_count += 1
            else:
                absent_count += 1

            row_data = [
                rec["roll_no"], rec["student_name"], rec["phone"], rec["email"],
                rec["entry_time"], rec["exit_time"], rec["duration_minutes"],
                rec["qr_verified"], rec["face_verified"], rec["status"]
            ]

            for c_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = thin_border
                if c_idx in [5, 6, 7, 8, 9, 10]:
                    cell.alignment = Alignment(horizontal="center")

                # Highlight Status
                if c_idx == 10:
                    if "PRESENT" in status:
                        cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
                        cell.font = Font(color="166534", bold=True)
                    else:
                        cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                        cell.font = Font(color="991B1B", bold=True)

        # Summary Row
        summary_row = start_row + len(records) + 2
        ws.cell(row=summary_row, column=1, value=f"Total Enrolled: {len(records)} | Present: {present_count} | Absent: {absent_count}").font = Font(bold=True)

        # Auto-fit columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


report_service = ReportService()
